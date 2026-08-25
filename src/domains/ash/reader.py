from __future__ import annotations

import re
from fnmatch import fnmatchcase
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from core.logging import log_file_read


class AshReader:
    def __init__(self, logger):
        self.logger = logger

    def read(self, file_path: str, cfg: dict[str, Any]) -> pd.DataFrame:
        log_file_read(self.logger, file_path, domain="ASH")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            reader_cfg = cfg.get("reader", {})
            column_patterns = self._column_patterns(reader_cfg.get("column_map", {}))
            required_columns = set(reader_cfg.get("required_columns", ["date"]))
            header_search_rows = int(reader_cfg.get("header_search_rows", 15))

            records: list[dict[str, Any]] = []
            used_sheets: set[str] = set()
            for material_key, material_cfg in cfg.get("materials", {}).items():
                sheet_name = self._resolve_sheet_name(
                    workbook.sheetnames,
                    material_cfg.get("sheet_patterns", []),
                    excluded=used_sheets,
                )
                if sheet_name is None:
                    self.logger.warning(
                        f"Ash sheet missing for {material_key}: "
                        f"{material_cfg.get('sheet_patterns', [])}"
                    )
                    continue

                used_sheets.add(sheet_name)
                worksheet = workbook[sheet_name]
                header_row, columns = self._find_header(
                    worksheet,
                    column_patterns,
                    required_columns,
                    header_search_rows,
                )
                if header_row is None:
                    self.logger.warning(
                        f"Ash header not found in sheet {sheet_name!r}; "
                        f"required columns={sorted(required_columns)}"
                    )
                    continue

                material_type = str(material_cfg["material_type"]).strip()
                self.logger.info(
                    f"Ash sheet matched: material_type={material_type!r}, "
                    f"sheet={sheet_name!r}, header_row={header_row}"
                )
                records.extend(
                    self._read_records(
                        worksheet,
                        header_row,
                        columns,
                        material_type,
                    )
                )

            return pd.DataFrame.from_records(records)
        finally:
            workbook.close()

    def _find_header(
        self,
        worksheet,
        column_patterns: dict[str, tuple[str, ...]],
        required_columns: set[str],
        search_rows: int,
    ) -> tuple[int | None, dict[str, int]]:
        best: tuple[int, int, dict[str, int]] | None = None
        max_row = min(max(1, search_rows), worksheet.max_row)

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=worksheet.max_column,
                values_only=True,
            ),
            start=1,
        ):
            matched_columns: dict[str, int] = {}
            for column_number, value in enumerate(values, start=1):
                field = self._match_field(value, column_patterns)
                if field and field not in matched_columns:
                    matched_columns[field] = column_number

            if not required_columns.issubset(matched_columns):
                continue

            candidate = (len(matched_columns), row_number, matched_columns)
            if best is None or candidate[0] > best[0]:
                best = candidate

        if best is None:
            return None, {}
        return best[1], best[2]

    @staticmethod
    def _read_records(
        worksheet,
        header_row: int,
        columns: dict[str, int],
        material_type: str,
    ) -> list[dict[str, Any]]:
        max_column = max(columns.values())
        records: list[dict[str, Any]] = []

        for values in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=1,
            max_col=max_column,
            values_only=True,
        ):
            record = {
                field: values[column_number - 1]
                for field, column_number in columns.items()
            }
            analysis_date = record.get("date")
            if analysis_date is None or not str(analysis_date).strip():
                continue

            if not any(
                value is not None and str(value).strip()
                for field, value in record.items()
                if field != "date"
            ):
                continue

            record["material_type"] = material_type
            records.append(record)

        return records

    def _column_patterns(
        self,
        column_map: dict[str, Any],
    ) -> dict[str, tuple[str, ...]]:
        patterns: dict[str, tuple[str, ...]] = {}
        for field, configured_patterns in column_map.items():
            if isinstance(configured_patterns, str):
                configured_patterns = [configured_patterns]
            normalized = tuple(
                pattern
                for value in configured_patterns or []
                if (pattern := self._normalize_pattern(value))
            )
            if normalized:
                patterns[str(field)] = normalized
        return patterns

    def _match_field(
        self,
        value: Any,
        column_patterns: dict[str, tuple[str, ...]],
    ) -> str | None:
        normalized = self._normalize(value)
        if not normalized:
            return None

        for field, patterns in column_patterns.items():
            if any(fnmatchcase(normalized, pattern) for pattern in patterns):
                return field
        return None

    def _resolve_sheet_name(
        self,
        sheet_names: list[str],
        configured_patterns: list[str] | str,
        excluded: set[str] | None = None,
    ) -> str | None:
        patterns = (
            [configured_patterns]
            if isinstance(configured_patterns, str)
            else list(configured_patterns)
        )
        normalized_patterns = [
            pattern
            for value in patterns
            if (pattern := self._normalize_pattern(value))
        ]
        excluded = excluded or set()

        for sheet_name in sheet_names:
            if sheet_name in excluded:
                continue
            normalized = self._normalize(sheet_name)
            if any(fnmatchcase(normalized, pattern) for pattern in normalized_patterns):
                return sheet_name
        return None

    @staticmethod
    def _normalize(value: Any) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())

    @staticmethod
    def _normalize_pattern(value: Any) -> str:
        return re.sub(r"[^a-z0-9*?]+", "", str(value or "").casefold())

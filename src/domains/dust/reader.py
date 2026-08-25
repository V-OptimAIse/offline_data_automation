from __future__ import annotations

import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from core.logging import log_file_read


class DustReader:
    def __init__(self, logger):
        self.logger = logger

    def read_basic(
        self,
        file_path: str,
        cfg: dict[str, Any],
        record_columns: dict[str, str],
    ) -> pd.DataFrame:
        log_file_read(self.logger, file_path, domain="DUST_BASIC")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            date_field = record_columns["date"]
            material_field = record_columns["material"]
            sheet_name = self._resolve_sheet_name(
                workbook.sheetnames,
                cfg["sheet_name"],
            )
            if sheet_name is None:
                self.logger.warning(f"Dust basic sheet missing: {cfg['sheet_name']}")
                return pd.DataFrame()

            worksheet = workbook[sheet_name]
            date_index = self._column_index(cfg["date_column"])
            material_specs = []
            max_column = date_index

            for material_cfg in cfg.get("materials", {}).values():
                field_columns = {
                    field: self._column_index(column)
                    for field, column in material_cfg.get("columns", {}).items()
                }
                if not field_columns:
                    continue
                material_specs.append(
                    (material_cfg["material_code"], field_columns)
                )
                max_column = max(max_column, *field_columns.values())

            records: list[dict[str, Any]] = []
            for values in worksheet.iter_rows(
                min_row=int(cfg.get("first_data_row", 1)),
                min_col=1,
                max_col=max_column,
                values_only=True,
            ):
                record_date = values[date_index - 1]
                if record_date is None or str(record_date).strip() == "":
                    continue

                for material_code, field_columns in material_specs:
                    record = {
                        date_field: record_date,
                        material_field: material_code,
                    }
                    for field, column_index in field_columns.items():
                        record[field] = values[column_index - 1]

                    if any(
                        value is not None and str(value).strip() != ""
                        for field, value in record.items()
                        if field not in {date_field, material_field}
                    ):
                        records.append(record)

            return pd.DataFrame.from_records(records)
        finally:
            workbook.close()

    def read_chemical(
        self,
        file_path: str,
        cfg: dict[str, Any],
        record_columns: dict[str, str],
    ) -> pd.DataFrame:
        log_file_read(self.logger, file_path, domain="DUST_CHEMICAL")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            date_field = record_columns["date"]
            material_field = record_columns["material"]
            min_column, max_column = self._column_bounds(cfg["columns"])
            header_row = int(cfg["header_row"])
            configured_map = {
                self._normalize_header(source): target
                for source, target in cfg.get("column_map", {}).items()
            }
            filter_cfg = cfg.get("row_filter") or {}
            filter_field = str(filter_cfg.get("field") or "").strip()
            filter_value = filter_cfg.get("equals")
            filter_normalizer = str(filter_cfg.get("normalizer") or "").strip()
            records: list[dict[str, Any]] = []

            for sheet_cfg in cfg.get("sheets", {}).values():
                sheet_name = self._resolve_sheet_name(
                    workbook.sheetnames,
                    sheet_cfg["sheet_name"],
                )
                if sheet_name is None:
                    self.logger.warning(
                        f"Dust chemical sheet missing: {sheet_cfg['sheet_name']}"
                    )
                    continue

                worksheet = workbook[sheet_name]
                header_values = next(
                    worksheet.iter_rows(
                        min_row=header_row,
                        max_row=header_row,
                        min_col=min_column,
                        max_col=max_column,
                        values_only=True,
                    ),
                    (),
                )
                if not header_values:
                    self.logger.warning(f"Dust chemical sheet empty: {sheet_name}")
                    continue
                targets_by_offset = {
                    offset: configured_map[normalized]
                    for offset, value in enumerate(header_values)
                    if (normalized := self._normalize_header(value)) in configured_map
                }

                required = {date_field}
                if filter_field:
                    required.add(filter_field)
                missing = required - set(targets_by_offset.values())
                if missing:
                    self.logger.warning(
                        f"Dust chemical sheet {sheet_name!r} missing columns: "
                        f"{sorted(missing)}"
                    )
                    continue

                for values in worksheet.iter_rows(
                    min_row=header_row + 1,
                    min_col=min_column,
                    max_col=max_column,
                    values_only=True,
                ):
                    record = {
                        target: values[offset]
                        for offset, target in targets_by_offset.items()
                    }
                    if filter_field and self._normalize_filter_value(
                        record.get(filter_field),
                        filter_normalizer,
                    ) != self._normalize_filter_value(
                        filter_value,
                        filter_normalizer,
                    ):
                        continue

                    if filter_field and filter_cfg.get("drop_after_filter", True):
                        record.pop(filter_field, None)
                    record[material_field] = sheet_cfg["material_code"]
                    records.append(record)

            return pd.DataFrame.from_records(records)
        finally:
            workbook.close()

    @staticmethod
    def _column_index(value: str | int) -> int:
        if isinstance(value, int):
            return value
        return column_index_from_string(str(value).strip())

    @classmethod
    def _column_bounds(cls, value: str) -> tuple[int, int]:
        start, end = [part.strip() for part in str(value).split(":", 1)]
        return cls._column_index(start), cls._column_index(end)

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())

    @staticmethod
    def _normalize_plant(value: Any) -> str:
        normalized = re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())
        match = re.fullmatch(r"BF0*(\d+)", normalized)
        if match:
            return f"BF{int(match.group(1))}"
        return normalized

    @classmethod
    def _normalize_filter_value(cls, value: Any, normalizer: str) -> str:
        if normalizer.casefold() == "plant":
            return cls._normalize_plant(value)
        return str(value or "").strip().casefold()

    @staticmethod
    def _normalize_sheet_name(value: str) -> str:
        return " ".join(str(value).casefold().split())

    def _resolve_sheet_name(
        self,
        sheet_names: list[str],
        configured_name: str,
    ) -> str | None:
        if configured_name in sheet_names:
            return configured_name

        configured = self._normalize_sheet_name(configured_name)
        for sheet_name in sheet_names:
            if self._normalize_sheet_name(sheet_name) == configured:
                return sheet_name
        return None

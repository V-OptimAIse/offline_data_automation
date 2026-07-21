import re
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook

from core.logging import log_file_read


class RMStockReader:
    def __init__(self, logger):
        self.logger = logger

    @staticmethod
    def _normalize_cell(value) -> str:
        if value is None or pd.isna(value):
            return ""
        return re.sub(r"\s+", " ", str(value)).strip().lower()

    @staticmethod
    def _excel_column_name(index: int) -> str:
        name = ""
        index += 1
        while index:
            index, remainder = divmod(index - 1, 26)
            name = f"{chr(65 + remainder)}{name}"
        return name

    def _find_header_row(self, df: pd.DataFrame) -> int:
        for row_idx, row in df.head(20).iterrows():
            labels = {self._normalize_cell(value) for value in row}
            if "particulars" in labels and "physical stock" in labels:
                return row_idx

        raise ValueError(
            "RM STOCK header row with PARTICULARS and Physical Stock not found"
        )

    def _find_column(self, df: pd.DataFrame, header_row: int, label: str) -> int:
        expected = self._normalize_cell(label)
        for col_idx, value in df.iloc[header_row].items():
            if self._normalize_cell(value) == expected:
                return col_idx

        available = [
            str(value).strip()
            for value in df.iloc[header_row].tolist()
            if value is not None and not pd.isna(value)
        ]
        raise ValueError(
            f"RM STOCK column '{label}' not found. Header labels: {available}"
        )

    def _find_data_end_row(self, df: pd.DataFrame, header_row: int) -> int:
        for row_idx, row in df.iloc[header_row + 1 :].iterrows():
            labels = {self._normalize_cell(value) for value in row}
            if "sinter plant" in labels:
                return row_idx

        return len(df)

    def _extract_timestamp(self, df: pd.DataFrame, run_date: str) -> pd.Timestamp:
        for _, row in df.head(8).iterrows():
            for value in row:
                ts = None
                if isinstance(value, pd.Timestamp):
                    ts = value
                elif isinstance(value, (datetime, date)):
                    ts = pd.Timestamp(value)
                elif isinstance(value, str) and value.strip():
                    ts = pd.to_datetime(value, errors="coerce", dayfirst=True)

                if ts is not None and pd.notna(ts) and 2000 <= ts.year <= 2100:
                    return pd.Timestamp(ts)

        self.logger.warning(
            f"RM STOCK timestamp not found in sheet header for {run_date}; "
            "using run date"
        )
        return pd.to_datetime(run_date, format="%d-%b-%Y")

    def _get_sheet_name(self, run_date: str) -> str:
        dt = pd.to_datetime(run_date, format="%d-%b-%Y")
        return dt.strftime("%d.%m")

    def _get_sinter_sheet_name(self, sheet_names: list[str], run_date: str) -> str:
        dt = pd.to_datetime(run_date, format="%d-%b-%Y")
        sept = dt.strftime("%b").upper().replace("SEP", "SEPT")
        candidates = [
            dt.strftime("%B-%Y").upper(),
            dt.strftime("%B %Y").upper(),
            f"{sept}-{dt.year}",
        ]
        sheets = {s.strip().upper(): s for s in sheet_names}
        for name in candidates:
            if name in sheets:
                return sheets[name]
        raise ValueError(f"Sinter DPR sheet not found for {run_date}")

    def read(self, file_path: str, run_date: str) -> tuple[pd.DataFrame, pd.Timestamp]:
        sheet_name = self._get_sheet_name(run_date)

        log_file_read(self.logger, file_path, domain="RM_STOCK", sheet=sheet_name)
        self.logger.info(f"Reading sheet: {sheet_name}")

        try:
            raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        except ValueError:
            raise ValueError(f"Sheet '{sheet_name}' not found in file")

        header_row = self._find_header_row(raw_df)
        material_col = self._find_column(raw_df, header_row, "PARTICULARS")
        stock_col = self._find_column(raw_df, header_row, "Physical Stock")
        data_end_row = self._find_data_end_row(raw_df, header_row)

        df = raw_df.iloc[header_row + 1:data_end_row, [material_col, stock_col]].copy()
        df.columns = ["material", "physical_stock"]

        df = df[df["material"].notna()].copy()
        df["material"] = df["material"].astype(str).str.strip()
        df = df[df["material"].ne("")].copy()
        df["physical_stock"] = pd.to_numeric(df["physical_stock"], errors="coerce")

        ts = self._extract_timestamp(raw_df, run_date)
        self.logger.info(
            "RM STOCK columns resolved: "
            f"material={self._excel_column_name(material_col)}, "
            f"physical_stock={self._excel_column_name(stock_col)}, "
            f"non_blank_stock_rows={int(df['physical_stock'].notna().sum())}"
        )

        return df, ts

    def read_sinter_stock(self, file_path: str, run_date: str) -> pd.DataFrame:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet_name = self._get_sinter_sheet_name(wb.sheetnames, run_date)
            rows = list(
                wb[sheet_name].iter_rows(
                    min_row=1,
                    max_row=80,
                    values_only=True,
                )
            )
        finally:
            wb.close()

        log_file_read(self.logger, file_path, domain="RM_STOCK_SINTER", sheet=sheet_name)
        self.logger.info(f"Reading sinter stock sheet: {sheet_name}")

        def clean(value) -> str:
            return str(value).strip().lower() if value is not None else ""

        stock_row = next(
            (idx for idx, row in enumerate(rows) if any(clean(v) == "stock qty" for v in row)),
            None,
        )
        if stock_row is None:
            raise ValueError(f"STOCK QTY row not found in sheet '{sheet_name}'")

        date_row = next(
            (idx for idx, row in enumerate(rows) if any(clean(v) == "date" for v in row)),
            4,
        )
        day = pd.to_datetime(run_date, format="%d-%b-%Y").day
        day_col = next(
            (
                idx
                for idx, value in enumerate(rows[date_row])
                if pd.to_numeric(value, errors="coerce") == day
            ),
            None,
        )
        if day_col is None:
            raise ValueError(f"Day {day} not found in sheet '{sheet_name}'")

        stock = pd.to_numeric(rows[stock_row][day_col], errors="coerce")
        if pd.isna(stock) or stock <= 0:
            self.logger.info(
                f"No positive sinter stock value for {run_date} in sheet '{sheet_name}'; "
                "skipping sinter stock row"
            )
            return pd.DataFrame(columns=["material", "physical_stock"])

        return pd.DataFrame(
            {"material": ["Sinter Stock at Yard"], "physical_stock": [stock]}
        )

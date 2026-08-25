from __future__ import annotations

import logging
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from core.config_loader import load_yaml
from app import _ash_source_files
from domains.ash.reader import AshReader
from domains.ash.service import AshService
from domains.ash.transformer import AshTransformer
from domains.download.service import PortalDownloader, format_portal_filename
from infrastructure.database_targets import DatabaseTarget


LOGGER = logging.getLogger("test_ash")
ASH_CONFIG = load_yaml("src/config/ash.yaml")["ash"]


class AshReaderTransformerTests(unittest.TestCase):
    def test_reader_discovers_shifted_headers_and_year_suffixed_sheets(self):
        workbook = Workbook()
        coke = workbook.active
        coke.title = "FY26-27 COKE PROD."
        nutcoke = workbook.create_sheet("BF-01 & 02 NUT COKE 26-27")
        pci = workbook.create_sheet("BF-01 & 02 PCI 26-27")

        self._write_material_table(
            coke,
            header_row=4,
            start_column=2,
            headers=self._headers("Coke", include_alkali_oxides=False),
            values=self._values("31.07.2026", include_alkali_oxides=False),
        )
        self._write_material_table(
            nutcoke,
            header_row=7,
            start_column=4,
            headers=self._headers("Nut Coke", include_alkali_oxides=False),
            values=self._values("31.07.2026", include_alkali_oxides=False),
        )
        self._write_material_table(
            pci,
            header_row=5,
            start_column=1,
            headers=self._headers("PCI Coal", include_alkali_oxides=True),
            values=self._values("31.07.26", include_alkali_oxides=True),
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "ASH ANALYSIS 26-27.xlsx"
            workbook.save(workbook_path)
            raw = AshReader(LOGGER).read(str(workbook_path), ASH_CONFIG)

        result = AshTransformer(LOGGER).transform(
            raw,
            ASH_CONFIG,
            [date(2026, 7, 31)],
        )

        self.assertEqual(result["material_type"].tolist(), ["coke", "nutcoke", "pci"])
        self.assertEqual(result["date"].tolist(), [date(2026, 7, 31)] * 3)
        self.assertTrue((result["sio2"] == 55.1).all())
        self.assertTrue(pd.isna(result.loc[result["material_type"] == "coke", "na2o"]).all())
        pci_row = result[result["material_type"] == "pci"].iloc[0]
        self.assertEqual(pci_row["na2o"], 0.55)
        self.assertEqual(pci_row["total_alkali_in_material"], 0.25)

    @staticmethod
    def _headers(material_label: str, include_alkali_oxides: bool) -> list[str]:
        headers = [
            "DATE",
            "%IM",
            "% ASH",
            "% VM",
            "% FC",
            "% SiO2",
            "% Al2O3",
            "% CaO",
            "% MgO",
            "% Fe2O3",
            f"% S in {material_label}",
            f"% P in {material_label}",
        ]
        if include_alkali_oxides:
            headers.extend(["% Na2O", "% K2O"])
        headers.extend([f"% Total Alkali in {material_label}", "% TiO2"])
        return headers

    @staticmethod
    def _values(date_value: str, include_alkali_oxides: bool) -> list[float | str]:
        values: list[float | str] = [
            date_value,
            0.4,
            12.0,
            1.0,
            86.6,
            55.1,
            27.2,
            3.2,
            1.2,
            7.4,
            0.73,
            0.04,
        ]
        if include_alkali_oxides:
            values.extend([0.55, 1.6])
        values.extend([0.25, 1.5])
        return values

    @staticmethod
    def _write_material_table(
        sheet,
        header_row: int,
        start_column: int,
        headers: list[str],
        values: list[float | str],
    ) -> None:
        for offset, header in enumerate(headers):
            sheet.cell(header_row, start_column + offset, header)
        for offset, value in enumerate(values):
            sheet.cell(header_row + 1, start_column + offset, value)


class AshDatabaseTests(unittest.TestCase):
    def test_database_sync_uses_material_and_date_as_upsert_key(self):
        frame = pd.DataFrame(
            [
                {
                    "material_type": "coke",
                    "date": date(2026, 7, 31),
                    "im": 0.46,
                    "ash": 12.36,
                }
            ]
        )
        calls = []

        class FakeClient:
            def fetch_table_columns(self, schema, table_names):
                return {
                    "ash_chemical_analysis": {
                        "material_type",
                        "date",
                        "im",
                        "ash",
                        "updated_at",
                    }
                }

            def insert_dataframe(self, **kwargs):
                calls.append(kwargs)
                return len(kwargs["df"])

        def fake_write_to_targets(setting_cfg, logger, domain, writer):
            writer(FakeClient(), DatabaseTarget("pi_db", "PI_DB", {}))

        with patch(
            "domains.ash.service.write_to_database_targets",
            side_effect=fake_write_to_targets,
        ):
            AshService(LOGGER)._push_to_database_targets(
                frame,
                {"write_db": ["pi_db"]},
                ASH_CONFIG,
            )

        self.assertEqual(len(calls), 1)
        self.assertEqual(calls[0]["table_name"], "offline_feed.ash_chemical_analysis")
        self.assertEqual(calls[0]["conflict_cols"], ["material_type", "date"])
        self.assertFalse(calls[0]["null_non_positive_values"])
        self.assertIn("updated_at", calls[0]["df"].columns)


class AshDownloadTests(unittest.TestCase):
    def test_financial_year_filename_template_changes_automatically(self):
        template = "ASH ANALYSIS {financial_year_short}"
        self.assertEqual(format_portal_filename(template, "25-Aug-2026"), "ASH ANALYSIS 26-27")
        self.assertEqual(format_portal_filename(template, "31-Mar-2027"), "ASH ANALYSIS 26-27")
        self.assertEqual(format_portal_filename(template, "01-Apr-2027"), "ASH ANALYSIS 27-28")

    def test_expected_financial_year_is_preferred_over_newer_generic_match(self):
        reader = PortalDownloader(None, None, LOGGER)
        rows = [
            {
                "name": "ASH ANALYSIS 27-28.xlsx",
                "modified": "08/25/2027 12:00:00",
            },
            {
                "name": "ASH ANALYSIS 26-27.xlsx",
                "modified": "08/25/2026 12:00:00",
            },
        ]

        match = reader._find_latest_matching_file(
            rows,
            ["ash", "analysis"],
            "ASH ANALYSIS 26-27",
        )

        self.assertEqual(match["name"], "ASH ANALYSIS 26-27.xlsx")

    def test_skip_download_ignores_excel_lock_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            download_dir = Path(temp_dir)
            source = download_dir / "ASH ANALYSIS 26-27.xlsx"
            lock_file = download_dir / "~$ASH ANALYSIS 26-27.xlsx"
            source.touch()
            lock_file.touch()

            matches = _ash_source_files(
                skip_download=True,
                download_dir=download_dir,
                download_result=None,
                run_dates=["31-Jul-2026"],
                filename_template="ASH ANALYSIS {financial_year_short}",
                logger=LOGGER,
            )

        self.assertEqual(matches, [source])


if __name__ == "__main__":
    unittest.main()

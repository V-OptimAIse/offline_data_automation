from __future__ import annotations

import logging
import sys
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from domains.dust.reader import DustReader
from domains.dust.service import DustService
from domains.dust.transformer import DustTransformer
from infrastructure.database_targets import DatabaseTarget


LOGGER = logging.getLogger("test_dust")
RECORD_COLUMNS = {
    "date": "date",
    "material": "material_code",
}


class DustReaderTransformerTests(unittest.TestCase):
    def setUp(self):
        self.reader = DustReader(LOGGER)
        self.transformer = DustTransformer(LOGGER)

    def test_basic_reader_extracts_configured_dust_blocks(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "GCP & ALL DUST"
        sheet.append(["DATE", None, "%Fe (T)", "%LOI", None, "%Ash", "%VM", "%FC", None, "%Fe (T)", "%LOI", None, "%Ash", "%VM", "%FC", "%Custom"])
        sheet.append([datetime(2026, 8, 5), None, 17.74, 54.19, None, 45.81, 22.24, 31.95, None, 35.64, 30.02, None, 69.98, 7.27, 22.75, 99.5])

        cfg = {
            "sheet_name": "GCP & ALL DUST",
            "date_column": "A",
            "first_data_row": 1,
            "materials": {
                "dust_catcher": {
                    "material_code": "dust_1",
                    "columns": {
                        "fe_total": "J",
                        "loi": "K",
                        "ash": "M",
                        "vm": "N",
                        "fc": "O",
                        "custom_metric": "P",
                    },
                },
                "gcp_dust": {
                    "material_code": "dust_2",
                    "columns": {
                        "fe_total": "C",
                        "loi": "D",
                        "ash": "F",
                        "vm": "G",
                        "fc": "H",
                    },
                },
            },
        }

        with patch("domains.dust.reader.load_workbook", return_value=workbook):
            raw = self.reader.read_basic("bunker.xlsx", cfg, RECORD_COLUMNS)
            result = self.transformer.transform(
                raw,
                "basic",
                [date(2026, 8, 5)],
                RECORD_COLUMNS,
            )

        self.assertEqual(result["material_code"].tolist(), ["dust_1", "dust_2"])
        dust_catcher = result[result["material_code"] == "dust_1"].iloc[0]
        self.assertEqual(dust_catcher["fe_total"], 35.64)
        self.assertEqual(dust_catcher["fc"], 22.75)
        self.assertEqual(dust_catcher["custom_metric"], 99.5)

    def test_chemical_reader_filters_bf2_and_averages_same_day_samples(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "DUST CATCHER"
        second = workbook.create_sheet("GCP DUST")
        headers = [
            None,
            "DoS",
            "XRF Application",
            "Sample name ",
            "Plant",
            "%Fe2O3",
            "%SiO2",
            "%Al2O3",
            "%TiO2",
            "%MnO",
            "%CaO",
            "%MgO",
            "%Na2O",
            "%K2O",
            "%P2O5",
            "%SO3",
            "%Cr2O3",
            "%ZnO",
            "%PbO",
            "%LoI",
        ]
        for sheet in (first, second):
            sheet.append([])
            sheet.append(headers)

        first.append([None, datetime(2026, 8, 19), None, "Dust Catcher", "BF-1", 10])
        first.append([None, datetime(2026, 8, 19), None, "Dust Catcher", "BF-02", 20])
        second.append([None, datetime(2026, 8, 19), None, "GCP Dust", "BF2", 30])
        second.append([None, datetime(2026, 8, 19), None, "GCP Dust", "BF-2", 40])

        cfg = {
            "row_filter": {
                "field": "plant",
                "equals": "BF2",
                "normalizer": "plant",
                "drop_after_filter": True,
            },
            "header_row": 2,
            "columns": "B:T",
            "column_map": {
                "DoS": "date",
                "Plant": "plant",
                "%Fe2O3": "fe2o3",
                "%SiO2": "sio2",
                "%Al2O3": "al2o3",
                "%TiO2": "tio2",
                "%MnO": "mno",
                "%CaO": "cao",
                "%MgO": "mgo",
                "%Na2O": "na2o",
                "%K2O": "k2o",
                "%P2O5": "p2o5",
                "%SO3": "so3",
                "%Cr2O3": "cr2o3",
                "%ZnO": "zno",
                "%PbO": "pbo",
                "%LoI": "loi",
            },
            "sheets": {
                "dust_catcher": {
                    "sheet_name": "DUST CATCHER",
                    "material_code": "dust_1",
                },
                "gcp_dust": {
                    "sheet_name": "GCP DUST",
                    "material_code": "dust_2",
                },
            },
        }

        with patch("domains.dust.reader.load_workbook", return_value=workbook):
            raw = self.reader.read_chemical(
                "dust_chemical.xlsx",
                cfg,
                RECORD_COLUMNS,
            )
            result = self.transformer.transform(
                raw,
                "chemical",
                [date(2026, 8, 19)],
                RECORD_COLUMNS,
            )

        self.assertEqual(len(raw), 3)
        self.assertEqual(len(result), 2)
        dust_catcher = result[result["material_code"] == "dust_1"].iloc[0]
        gcp_dust = result[result["material_code"] == "dust_2"].iloc[0]
        self.assertEqual(dust_catcher["fe2o3"], 20)
        self.assertEqual(gcp_dust["fe2o3"], 35)
        self.assertTrue(pd.isna(gcp_dust["sio2"]))

    def test_database_sync_writes_both_tables_to_both_postgres_targets(self):
        frames = {
            "basic": pd.DataFrame(
                [
                    {
                        "material_code": "dust_1",
                        "date": date(2026, 8, 19),
                        "fe_total": 0.0,
                    }
                ]
            ),
            "chemical": pd.DataFrame(
                [
                    {
                        "material_code": "dust_2",
                        "date": date(2026, 8, 19),
                        "fe2o3": 27.96,
                    }
                ]
            ),
        }
        dust_cfg = {
            "record_columns": RECORD_COLUMNS,
            "postgres": {
                "schema": "offline_feed",
                "upsert_mode": "update_insert",
                "conflict_cols": ["material_code", "date"],
                "tables": {
                    "basic": "dust_basic_analysis",
                    "chemical": "dust_chemical_analysis",
                },
                "material_master": {},
            }
        }
        calls = []

        class FakeClient:
            def fetch_material_codes(self, **kwargs):
                return {"dust_1", "dust_2"}

            def fetch_table_columns(self, schema, table_names):
                return {
                    "dust_basic_analysis": {
                        "material_code",
                        "date",
                        "fe_total",
                        "updated_at",
                    },
                    "dust_chemical_analysis": {
                        "material_code",
                        "date",
                        "fe2o3",
                        "updated_at",
                    },
                }

            def insert_dataframe(self, **kwargs):
                calls.append(kwargs)
                return len(kwargs["df"])

        def fake_write_to_targets(setting_cfg, logger, domain, writer):
            for key, label in (
                ("neon_developer", "NeonDB developer"),
                ("pi_db", "PI_DB"),
            ):
                writer(FakeClient(), DatabaseTarget(key, label, {}))

        with patch(
            "domains.dust.service.write_to_database_targets",
            side_effect=fake_write_to_targets,
        ):
            DustService(LOGGER)._push_to_database_targets(
                frames,
                {"write_db": ["neon_db", "pi_db"]},
                dust_cfg,
            )

        self.assertEqual(len(calls), 4)
        self.assertEqual(
            {call["table_name"] for call in calls},
            {
                "offline_feed.dust_basic_analysis",
                "offline_feed.dust_chemical_analysis",
            },
        )
        self.assertTrue(
            all(call["null_non_positive_values"] is False for call in calls)
        )
        basic_calls = [
            call for call in calls if call["table_name"].endswith("dust_basic_analysis")
        ]
        self.assertTrue(all(call["df"].iloc[0]["fe_total"] == 0 for call in basic_calls))


if __name__ == "__main__":
    unittest.main()
